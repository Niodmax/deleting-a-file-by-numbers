import os
import re
import pickle
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, Menu
import winshell

# Определение абсолютного пути для файла deleted_files.pickle
pickle_path = os.path.abspath("deleted_files.pickle")

def process_files_confirmation():
    file_path = file_entry.get()
    folder_path = folder_entry.get()

    if not file_path or not folder_path:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите файл и папку")
        return

    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
    except FileNotFoundError:
        messagebox.showerror("Ошибка", "Файл Excel не найден")
        return

    number_pattern = re.compile(r'\d+')
    numbers_from_excel = []

    for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
        for cell_value in row:
            cell_value_str = str(cell_value)
            numbers = number_pattern.findall(cell_value_str)
            numbers_from_excel.extend(numbers)

    deleted_files = []

    with os.scandir(folder_path) as entries:
        for entry in entries:
            if entry.is_file():
                filename = entry.name
                full_path = entry.path

                numbers_in_filename = number_pattern.findall(filename)
                numbers_in_filename = [str(number) for number in numbers_in_filename]

                should_delete = not any(number in numbers_from_excel for number in numbers_in_filename)

                if should_delete:
                    deleted_files.append((filename, full_path))

    if deleted_files:
        confirmation_message = f"Вы действительно хотите удалить {len(deleted_files)} файлов?"
        confirm_result = messagebox.askquestion("Подтверждение", confirmation_message)
        if confirm_result == "yes":
            actions_text = ""
            for filename, full_path in deleted_files:
                winshell.delete_file(os.path.abspath(full_path), no_confirm=True)
                actions_text += f"Перемещен в корзину: {filename}\n"

            actions_text += "Обработка завершена."
            actions_textbox.config(state=tk.NORMAL)
            actions_textbox.delete("1.0", tk.END)
            actions_textbox.insert(tk.END, actions_text)
            actions_textbox.config(state=tk.DISABLED)

            with open(pickle_path, "wb") as file:
                pickle.dump(deleted_files, file)
    else:
        messagebox.showinfo("Информация", "Нет файлов для удаления.")

def restore_files_from_trash():
    folder_path = folder_entry.get()
    deleted_files = []

    try:
        with open(pickle_path, "rb") as file:
            deleted_files = pickle.load(file)
            for filename, full_path in deleted_files:
                winshell.undelete(os.path.abspath(full_path))
                actions_text = f"Восстановлен из корзины: {filename}\n"
                actions_textbox.config(state=tk.NORMAL)
                actions_textbox.insert(tk.END, actions_text)
                actions_textbox.config(state=tk.DISABLED)
        os.remove(pickle_path)
    except FileNotFoundError:
        messagebox.showinfo("Информация", "Нет данных для восстановления файлов из корзины.")

def choose_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file_path)
    save_settings()

def choose_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)
    save_settings()

def save_settings():
    settings = {
        "file_path": file_entry.get(),
        "folder_path": folder_entry.get()
    }
    with open("settings.pickle", "wb") as file:
        pickle.dump(settings, file)

def load_settings():
    try:
        with open("settings.pickle", "rb") as file:
            settings = pickle.load(file)
            file_entry.insert(0, settings["file_path"])
            folder_entry.insert(0, settings["folder_path"])
    except FileNotFoundError:
        pass

# Создаем главное окно
root = tk.Tk()
root.title("Удаление файлов")

# Создаем и располагаем виджеты
file_label = tk.Label(root, text="Удаление производится если в таблице нет номера. Если таблица пустая удаляется всё содержимое в папке")
file_label.pack()

file_label = tk.Label(root, text="Выберите файл Excel:")
file_label.pack()

file_entry = tk.Entry(root)
file_entry.pack(fill="x", padx=10, pady=(0, 5)) # Здесь padx и pady - опциональные отступы

file_button = tk.Button(root, text="Выбрать файл", command=choose_file)
file_button.pack()

folder_label = tk.Label(root, text="Выберите папку для удаления:")
folder_label.pack()

folder_entry = tk.Entry(root)
folder_entry.pack(fill="x", padx=10, pady=(0, 5)) # Здесь padx и pady - опциональные отступы

folder_button = tk.Button(root, text="Выбрать папку", command=choose_folder)
folder_button.pack()

# Создание кнопки с подтверждением
process_button = tk.Button(root, text="Удалить файлы", command=process_files_confirmation, bg="orange", height=2)
process_button.place(x=80, y=135, width=120, height=30)

restore_button = tk.Button(root, text="Восстановить файлы из корзины", command=restore_files_from_trash, bg="green", height=2)
restore_button.place(x=420, y=135, width=220, height=30)

actions_label = tk.Label(root, text="Проделанные действия:")
actions_label.pack()

actions_textbox = scrolledtext.ScrolledText(root, wrap=tk.WORD, height=10, state=tk.DISABLED)
actions_textbox.pack()


def copy_text():
    actions_textbox.config(state=tk.NORMAL)
    selected_text = actions_textbox.selection_get()
    root.clipboard_clear()
    root.clipboard_append(selected_text)
    actions_textbox.config(state=tk.DISABLED)
    selected_text = actions_textbox.selection_get()
    root.clipboard_clear()
    root.clipboard_append(selected_text)

actions_textbox.bind("<Control-c>", lambda e: copy_text())
context_menu = Menu(actions_textbox, tearoff=0)
context_menu.add_command(label="Копировать", command=copy_text)

def show_context_menu(event):
    context_menu.tk_popup(event.x_root, event.y_root)

actions_textbox.bind("<Button-3>", show_context_menu)

# Загрузка сохраненных настроек
load_settings()

# Запускаем главный цикл
root.mainloop()
